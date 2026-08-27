from __future__ import annotations

from copy import copy
from io import BytesIO

import openpyxl
import pandas as pd

from matcher import normalize


def read_first_sheet(uploaded_file) -> tuple[pd.DataFrame, str]:
    """Read the first worksheet from an uploaded Excel file."""

    uploaded_file.seek(0)
    excel_file = pd.ExcelFile(uploaded_file)
    sheet_name = excel_file.sheet_names[0]
    uploaded_file.seek(0)
    dataframe = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    return dataframe, sheet_name


def get_header_columns(worksheet) -> dict[str, int]:
    """Return worksheet headers and their column numbers."""

    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }


def copy_cell_format(source_cell, target_cell) -> None:
    """Copy formatting from one Excel cell to another."""

    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def ensure_column_after(
    worksheet,
    previous_column_number: int,
    column_name: str,
    minimum_width: int = 18,
) -> int:
    """Find a column or create it after the specified column."""

    header_columns = get_header_columns(worksheet)

    if column_name in header_columns:
        return header_columns[column_name]

    new_column_number = previous_column_number + 1
    worksheet.insert_cols(new_column_number, amount=1)

    source_header = worksheet.cell(
        row=1,
        column=previous_column_number,
    )
    new_header = worksheet.cell(
        row=1,
        column=new_column_number,
    )
    new_header.value = column_name
    copy_cell_format(source_header, new_header)

    source_letter = openpyxl.utils.get_column_letter(
        previous_column_number
    )
    new_letter = openpyxl.utils.get_column_letter(
        new_column_number
    )
    source_width = worksheet.column_dimensions[source_letter].width
    worksheet.column_dimensions[new_letter].width = max(
        source_width or 12,
        minimum_width,
    )

    return new_column_number


def write_match_results(
    original_file_bytes: bytes,
    sheet_name: str,
    input_name_column: str,
    output_id_column: str,
    results: dict[str, dict],
    confidence_column_name: str = "Confidence Score",
    status_column_name: str = "Match Status",
) -> bytes:
    """Write College ID, status and confidence into the original file.

    Original row order, duplicate rows, existing formatting and all source
    values are preserved.
    """

    workbook = openpyxl.load_workbook(BytesIO(original_file_bytes))

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")

    worksheet = workbook[sheet_name]
    header_columns = get_header_columns(worksheet)

    if input_name_column not in header_columns:
        raise ValueError(
            f"College-name column not found: {input_name_column}"
        )

    if output_id_column not in header_columns:
        raise ValueError(
            f"College-ID column not found: {output_id_column}"
        )

    name_column_number = header_columns[input_name_column]
    id_column_number = header_columns[output_id_column]

    status_column_number = ensure_column_after(
        worksheet,
        id_column_number,
        status_column_name,
        minimum_width=18,
    )

    # Re-read headings because inserting Match Status can shift an existing
    # Confidence Score column.
    header_columns = get_header_columns(worksheet)
    id_column_number = header_columns[output_id_column]
    status_column_number = header_columns[status_column_name]

    confidence_column_number = ensure_column_after(
        worksheet,
        status_column_number,
        confidence_column_name,
        minimum_width=18,
    )

    header_columns = get_header_columns(worksheet)
    name_column_number = header_columns[input_name_column]
    id_column_number = header_columns[output_id_column]
    status_column_number = header_columns[status_column_name]
    confidence_column_number = header_columns[confidence_column_name]

    for row_number in range(2, worksheet.max_row + 1):
        original_college_name = worksheet.cell(
            row=row_number,
            column=name_column_number,
        ).value
        normalized_name = normalize(original_college_name)

        if normalized_name not in results:
            continue

        result = results[normalized_name]
        decision = str(result.get("decision", "NOT_FOUND"))
        college_id_value = result.get("college_id", "Not Found")
        confidence_value = float(result.get("confidence", 0.0) or 0.0)

        if decision == "NOT_FOUND":
            college_id_value = "Not Found"
            displayed_status = "NOT FOUND"
        elif decision == "NEEDS_REVIEW":
            college_id_value = "Needs Review"
            displayed_status = "NEEDS REVIEW"
        else:
            displayed_status = "FOUND"

        id_cell = worksheet.cell(row=row_number, column=id_column_number)
        status_cell = worksheet.cell(
            row=row_number,
            column=status_column_number,
        )
        confidence_cell = worksheet.cell(
            row=row_number,
            column=confidence_column_number,
        )

        id_cell.value = college_id_value
        status_cell.value = displayed_status
        confidence_cell.value = confidence_value

        copy_cell_format(id_cell, status_cell)
        copy_cell_format(id_cell, confidence_cell)
        confidence_cell.number_format = "0.00"

    output_file = BytesIO()
    workbook.save(output_file)
    return output_file.getvalue()